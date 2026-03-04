from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from queue import Empty, Queue
import subprocess
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


RAW_SHEET = "RAW"
COVER_SHEET = "cover"
DATE_CELL = "H7"
FONT_CANDIDATES = [
    Path("/System/Library/Fonts/Supplemental/AppleGothic.ttf"),
    Path("/System/Library/Fonts/AppleSDGothicNeo.ttc"),
]
DEFAULT_OUTPUT_DIR = Path.home() / "Desktop" / "JEBON_OUTPUT"


@dataclass
class CoverRecord:
    volume: str
    work_date: str
    payment_no: str


def safe_filename_part(value: object) -> str:
    text = str(value).strip()
    if not text:
        return "unknown"
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, "-")
    return " ".join(text.split())


def normalize_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_volume_spec(text: str) -> list[str]:
    normalized = text.replace("\n", ",").replace("~", "-")
    tokens = [tok.strip() for tok in normalized.split(",") if tok.strip()]
    if not tokens:
        raise ValueError("권 범위가 비어 있습니다. 예: 1-100 또는 1,3,5-8")

    result: list[str] = []
    seen: set[str] = set()

    for token in tokens:
        if "-" in token:
            parts = [p.strip() for p in token.split("-", 1)]
            if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
                raise ValueError(f"권 범위 형식 오류: '{token}'")

            start = int(parts[0])
            end = int(parts[1])
            step = 1 if start <= end else -1
            for num in range(start, end + step, step):
                vol = str(num)
                if vol not in seen:
                    seen.add(vol)
                    result.append(vol)
        else:
            vol = str(int(token)) if token.isdigit() else token
            if vol not in seen:
                seen.add(vol)
                result.append(vol)

    if not result:
        raise ValueError("유효한 권 정보가 없습니다.")
    return result


def parse_records_from_text(text: str, default_date: str = "") -> list[CoverRecord]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        raise ValueError("붙여넣기 데이터가 비어 있습니다.")

    records: list[CoverRecord] = []
    for idx, line in enumerate(lines):
        if "\t" in line:
            cols = [col.strip() for col in line.split("\t")]
            if len(cols) < 1:
                continue
            volume = cols[0] if len(cols) >= 1 else ""
            work_date = cols[1] if len(cols) >= 2 else ""
            payment_no = cols[2] if len(cols) >= 3 else ""
        else:
            cols = [col.strip() for col in line.split(",", 2)]
            volume = cols[0] if len(cols) >= 1 else ""
            work_date = cols[1] if len(cols) >= 2 else ""
            payment_no = cols[2] if len(cols) >= 3 else ""

        if idx == 0 and volume in {"권", "volume", "Volume"}:
            continue
        if not volume:
            continue

        if not work_date:
            work_date = default_date

        records.append(CoverRecord(volume=volume, work_date=work_date, payment_no=payment_no))

    if not records:
        raise ValueError("유효한 레코드를 읽지 못했습니다. (권/날짜/지급번호 3열 복사 필요)")
    return records


def read_default_date_string(excel_path: Path) -> str:
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
    except FileNotFoundError:
        raise
    except Exception as exc:
        raise RuntimeError(f"Failed to open '{excel_path.name}': {exc}") from exc

    try:
        ws = wb[COVER_SHEET]
    except KeyError as exc:
        wb.close()
        raise KeyError(f"Sheet '{COVER_SHEET}' not found in '{excel_path.name}'.") from exc

    raw_value = ws[DATE_CELL].value
    epoch = wb.epoch
    wb.close()

    if raw_value is None:
        return ""

    if isinstance(raw_value, (datetime, date)):
        return raw_value.strftime("%Y-%m-%d")

    if isinstance(raw_value, (int, float)):
        try:
            converted = from_excel(raw_value, epoch=epoch)
            if isinstance(converted, datetime):
                return converted.strftime("%Y-%m-%d")
            if isinstance(converted, date):
                return converted.strftime("%Y-%m-%d")
        except Exception:
            pass

    return str(raw_value).strip()


def read_cover_records(excel_path: Path, default_date: str = "") -> list[CoverRecord]:
    try:
        df = pd.read_excel(
            excel_path,
            sheet_name=RAW_SHEET,
            usecols="A:C",
            header=0,
            engine="openpyxl",
        )
    except ValueError as exc:
        raise ValueError(f"Sheet '{RAW_SHEET}' not found in '{excel_path.name}'.") from exc
    except Exception as exc:
        raise RuntimeError(f"Failed to read data from '{excel_path.name}': {exc}") from exc

    if df.empty:
        raise ValueError(f"No data found in '{RAW_SHEET}'.")

    records: list[CoverRecord] = []
    for _, row in df.iterrows():
        volume_text = normalize_cell_text(row.iloc[0])
        if not volume_text:
            continue

        date_text = normalize_cell_text(row.iloc[1]) if len(row) >= 2 else ""
        payment_text = normalize_cell_text(row.iloc[2]) if len(row) >= 3 else ""

        if not date_text:
            date_text = default_date

        records.append(CoverRecord(volume=volume_text, work_date=date_text, payment_no=payment_text))

    if not records:
        raise ValueError(f"No valid volume values found in '{RAW_SHEET}' column A (from row 2).")

    return records


def add_korean_font(pdf: FPDF) -> tuple[str, str]:
    errors: list[str] = []

    for idx, font_path in enumerate(FONT_CANDIDATES, start=1):
        if not font_path.exists():
            errors.append(f"{font_path} (not found)")
            continue

        family = f"KoreanFont{idx}"
        try:
            pdf.add_font(family=family, fname=str(font_path))
            return family, str(font_path)
        except TypeError:
            try:
                pdf.add_font(family, "", str(font_path), uni=True)
                return family, str(font_path)
            except Exception as exc:
                errors.append(f"{font_path} ({exc})")
        except Exception as exc:
            errors.append(f"{font_path} ({exc})")

    raise RuntimeError("No usable Korean font found. Tried: " + " | ".join(errors))


def generate_cover_pdf(record: CoverRecord, output_dir: Path) -> Path:
    if not record.work_date:
        raise ValueError(f"Volume {record.volume}: 작업일이 비어 있습니다.")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    font_family, _ = add_korean_font(pdf)
    page_width = 210

    # Top: fixed template title
    pdf.set_font(font_family, size=36)
    pdf.set_xy(0, 28)
    pdf.cell(page_width, 20, "제 본 표 지", align="C")

    # Middle: volume
    pdf.set_font(font_family, size=64)
    pdf.set_xy(0, 120)
    pdf.cell(page_width, 32, f"제 {record.volume} 권", align="C")

    # Bottom: payment + date
    payment_label = record.payment_no if record.payment_no else "-"
    pdf.set_font(font_family, size=18)
    pdf.set_xy(0, 248)
    pdf.cell(page_width, 12, f"지급명령번호: {payment_label}", align="C")

    pdf.set_font(font_family, size=20)
    pdf.set_xy(0, 263)
    pdf.cell(page_width, 12, f"작업일: {record.work_date}", align="C")

    output_name = f"{safe_filename_part(record.work_date)}_cover_{safe_filename_part(record.volume)}.pdf"
    output_path = output_dir / output_name
    pdf.output(str(output_path))
    return output_path


class CoverGeneratorApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("JEBON Cover Generator")
        self.root.geometry("1120x760")
        self.root.minsize(980, 640)

        self.log_queue: Queue[tuple[str, object]] = Queue()
        self.cancel_event = threading.Event()
        self.worker_running = False
        self.worker_thread: threading.Thread | None = None

        self.preview_records: list[CoverRecord] = []
        self.start_time: float | None = None

        default_excel = Path(__file__).resolve().parent / "data.xlsx"
        self.excel_path_var = tk.StringVar(value=str(default_excel))
        self.output_dir_var = tk.StringVar(value=str(DEFAULT_OUTPUT_DIR))

        self.manual_date_var = tk.StringVar(value="")
        self.manual_payment_var = tk.StringVar(value="")
        self.volume_spec_var = tk.StringVar(value="1-100")
        self.apply_manual_var = tk.BooleanVar(value=True)

        self.status_var = tk.StringVar(value="1) 파일 선택/클립보드/권범위 중 하나로 입력 → 2) 확인 → 3) 생성 실행")
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_text_var = tk.StringVar(value="0 / 0 (0%)")

        self.date_var = tk.StringVar(value="-")
        self.total_var = tk.StringVar(value="0")
        self.done_var = tk.StringVar(value="0")
        self.failed_var = tk.StringVar(value="0")
        self.remaining_var = tk.StringVar(value="0")
        self.current_var = tk.StringVar(value="-")
        self.elapsed_var = tk.StringVar(value="00:00")

        self._build_style()
        self._build_ui()
        self._poll_queue()

    def _build_style(self) -> None:
        style = ttk.Style(self.root)
        style.configure("Header.TLabel", font=("Apple SD Gothic Neo", 18, "bold"))
        style.configure("SubHeader.TLabel", font=("Apple SD Gothic Neo", 11))
        style.configure("CardTitle.TLabel", font=("Apple SD Gothic Neo", 10, "bold"))
        style.configure("CardValue.TLabel", font=("SF Pro Text", 14, "bold"))

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=14)
        main.pack(fill="both", expand=True)

        header = ttk.Frame(main)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="JEBON PDF Cover Generator", style="Header.TLabel").pack(anchor="w")
        ttk.Label(
            header,
            text="표지 양식 고정 + 날짜/지급명령번호 입력 + 엑셀/클립보드/권범위 일괄입력",
            style="SubHeader.TLabel",
        ).pack(anchor="w", pady=(2, 0))

        io_frame = ttk.LabelFrame(main, text="입력 / 출력 경로", padding=10)
        io_frame.pack(fill="x")

        ttk.Label(io_frame, text="Excel 파일").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=4)
        self.excel_entry = ttk.Entry(io_frame, textvariable=self.excel_path_var)
        self.excel_entry.grid(row=0, column=1, sticky="ew", pady=4)
        self.excel_browse_btn = ttk.Button(io_frame, text="찾기", command=self._browse_excel)
        self.excel_browse_btn.grid(row=0, column=2, padx=(8, 0), pady=4)

        ttk.Label(io_frame, text="출력 폴더").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=4)
        self.output_entry = ttk.Entry(io_frame, textvariable=self.output_dir_var)
        self.output_entry.grid(row=1, column=1, sticky="ew", pady=4)
        self.output_browse_btn = ttk.Button(io_frame, text="찾기", command=self._browse_output)
        self.output_browse_btn.grid(row=1, column=2, padx=(8, 0), pady=4)
        io_frame.grid_columnconfigure(1, weight=1)

        manual_frame = ttk.LabelFrame(main, text="빠른 입력 (엑셀 값 덮어쓰기)", padding=10)
        manual_frame.pack(fill="x", pady=(10, 0))

        ttk.Label(manual_frame, text="작업일").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        self.manual_date_entry = ttk.Entry(manual_frame, textvariable=self.manual_date_var, width=26)
        self.manual_date_entry.grid(row=0, column=1, sticky="w", pady=4)

        ttk.Label(manual_frame, text="지급명령번호").grid(row=0, column=2, sticky="w", padx=(18, 8), pady=4)
        self.manual_payment_entry = ttk.Entry(manual_frame, textvariable=self.manual_payment_var, width=32)
        self.manual_payment_entry.grid(row=0, column=3, sticky="w", pady=4)

        self.apply_manual_chk = ttk.Checkbutton(
            manual_frame,
            text="생성 시 위 입력값으로 덮어쓰기",
            variable=self.apply_manual_var,
        )
        self.apply_manual_chk.grid(row=0, column=4, sticky="w", padx=(16, 8), pady=4)

        self.apply_manual_btn = ttk.Button(manual_frame, text="미리보기에 즉시 반영", command=self._apply_manual_to_preview)
        self.apply_manual_btn.grid(row=0, column=5, sticky="e", padx=(8, 0), pady=4)

        ttk.Label(manual_frame, text="권 범위").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(8, 4))
        self.volume_spec_entry = ttk.Entry(manual_frame, textvariable=self.volume_spec_var, width=26)
        self.volume_spec_entry.grid(row=1, column=1, sticky="w", pady=(8, 4))

        ttk.Label(
            manual_frame,
            text="예: 1-100, 1~20, 1,3,5-8",
            style="SubHeader.TLabel",
        ).grid(row=1, column=2, columnspan=2, sticky="w", pady=(8, 4))

        self.clipboard_load_btn = ttk.Button(
            manual_frame,
            text="클립보드 일괄입력",
            command=self._load_records_from_clipboard,
        )
        self.clipboard_load_btn.grid(row=1, column=4, sticky="e", padx=(8, 0), pady=(8, 4))

        self.range_build_btn = ttk.Button(manual_frame, text="권 자동생성", command=self._build_preview_from_range)
        self.range_build_btn.grid(row=1, column=5, sticky="e", padx=(8, 0), pady=(8, 4))

        manual_frame.grid_columnconfigure(6, weight=1)

        action_frame = ttk.Frame(main)
        action_frame.pack(fill="x", pady=(10, 10))

        self.validate_btn = ttk.Button(action_frame, text="데이터 확인", command=self._validate_data)
        self.validate_btn.pack(side="left")

        self.generate_btn = ttk.Button(action_frame, text="PDF 생성 시작", command=self._start_generation, state="disabled")
        self.generate_btn.pack(side="left", padx=(8, 0))

        self.cancel_btn = ttk.Button(action_frame, text="중단", command=self._cancel_generation, state="disabled")
        self.cancel_btn.pack(side="left", padx=(8, 0))

        self.open_output_btn = ttk.Button(action_frame, text="출력 폴더 열기", command=self._open_output_folder)
        self.open_output_btn.pack(side="left", padx=(8, 0))

        dashboard = ttk.LabelFrame(main, text="작업 대시보드", padding=10)
        dashboard.pack(fill="x")

        cards = [
            ("작업일", self.date_var),
            ("전체", self.total_var),
            ("완료", self.done_var),
            ("실패", self.failed_var),
            ("남음", self.remaining_var),
            ("현재 권", self.current_var),
            ("경과시간", self.elapsed_var),
        ]

        for idx, (title, var) in enumerate(cards):
            card = ttk.Frame(dashboard, padding=(6, 4))
            card.grid(row=0, column=idx, sticky="nsew")
            ttk.Label(card, text=title, style="CardTitle.TLabel").pack(anchor="w")
            ttk.Label(card, textvariable=var, style="CardValue.TLabel").pack(anchor="w", pady=(4, 0))
            dashboard.grid_columnconfigure(idx, weight=1)

        progress_frame = ttk.LabelFrame(main, text="진행률", padding=10)
        progress_frame.pack(fill="x", pady=(10, 10))

        self.progress = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate", maximum=100, variable=self.progress_var)
        self.progress.pack(fill="x")
        ttk.Label(progress_frame, textvariable=self.progress_text_var).pack(anchor="w", pady=(6, 0))
        ttk.Label(progress_frame, textvariable=self.status_var).pack(anchor="w", pady=(2, 0))

        content = ttk.PanedWindow(main, orient="vertical")
        content.pack(fill="both", expand=True)

        table_frame = ttk.LabelFrame(content, text="권별 처리 상태", padding=8)
        self.volume_table = ttk.Treeview(
            table_frame,
            columns=("idx", "volume", "work_date", "payment_no", "status", "detail"),
            show="headings",
            height=12,
        )
        self.volume_table.heading("idx", text="No")
        self.volume_table.heading("volume", text="권")
        self.volume_table.heading("work_date", text="작업일")
        self.volume_table.heading("payment_no", text="지급명령번호")
        self.volume_table.heading("status", text="상태")
        self.volume_table.heading("detail", text="결과/오류")

        self.volume_table.column("idx", width=60, anchor="center", stretch=False)
        self.volume_table.column("volume", width=80, anchor="center", stretch=False)
        self.volume_table.column("work_date", width=180, anchor="center", stretch=False)
        self.volume_table.column("payment_no", width=280, anchor="w", stretch=False)
        self.volume_table.column("status", width=90, anchor="center", stretch=False)
        self.volume_table.column("detail", width=560, anchor="w", stretch=True)

        self.volume_table.tag_configure("waiting", foreground="#7f8c8d")
        self.volume_table.tag_configure("running", foreground="#0b69ff")
        self.volume_table.tag_configure("done", foreground="#1e8449")
        self.volume_table.tag_configure("failed", foreground="#c0392b")

        table_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.volume_table.yview)
        self.volume_table.configure(yscrollcommand=table_scroll.set)
        self.volume_table.pack(side="left", fill="both", expand=True)
        table_scroll.pack(side="right", fill="y")

        log_frame = ttk.LabelFrame(content, text="실시간 로그", padding=8)
        self.log_text = tk.Text(log_frame, wrap="word", state="disabled", height=8)
        log_scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        log_scroll.pack(side="right", fill="y")

        content.add(table_frame, weight=3)
        content.add(log_frame, weight=2)

    def _browse_excel(self) -> None:
        selected = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All Files", "*.*")],
        )
        if selected:
            self.excel_path_var.set(selected)

    def _browse_output(self) -> None:
        selected = filedialog.askdirectory(title="출력 폴더 선택")
        if selected:
            self.output_dir_var.set(selected)

    def _open_output_folder(self) -> None:
        output_dir = Path(self.output_dir_var.get().strip()).expanduser()
        if not output_dir.exists():
            messagebox.showwarning("폴더 없음", f"출력 폴더가 없습니다:\n{output_dir}")
            return
        try:
            subprocess.run(["open", str(output_dir)], check=False)
        except Exception as exc:
            messagebox.showerror("열기 실패", f"폴더 열기에 실패했습니다:\n{exc}")

    def _load_records_from_clipboard(self) -> None:
        if self.worker_running:
            return

        try:
            text = self.root.clipboard_get()
        except Exception:
            messagebox.showerror("클립보드 오류", "클립보드에서 텍스트를 읽을 수 없습니다.")
            return

        try:
            default_date = self.manual_date_var.get().strip()
            records = parse_records_from_text(text, default_date=default_date)
            self.preview_records = records

            self._refresh_table(self.preview_records)
            self._refresh_summary_metrics()
            self.progress_var.set(0)
            self.progress_text_var.set(f"0 / {len(self.preview_records)} (0%)")
            self.status_var.set("클립보드 데이터를 불러왔습니다. 검토 후 생성하세요.")
            self.generate_btn.configure(state="normal")

            self._append_log("--- 클립보드 일괄입력 완료 ---")
            self._append_log(f"읽은 권수: {len(self.preview_records)}")
            self._append_log("형식: 권[TAB]날짜[TAB]지급명령번호")
        except Exception as exc:
            self.generate_btn.configure(state="disabled")
            self._append_log(f"Error: {exc}")
            messagebox.showerror("클립보드 일괄입력 실패", str(exc))

    def _build_preview_from_range(self) -> None:
        if self.worker_running:
            return

        try:
            volumes = parse_volume_spec(self.volume_spec_var.get().strip())
            manual_date = self.manual_date_var.get().strip()
            manual_payment = self.manual_payment_var.get().strip()

            self.preview_records = [
                CoverRecord(volume=v, work_date=manual_date, payment_no=manual_payment)
                for v in volumes
            ]

            self._refresh_table(self.preview_records)
            self._refresh_summary_metrics()
            self.progress_var.set(0)
            self.progress_text_var.set(f"0 / {len(self.preview_records)} (0%)")
            self.status_var.set("권 범위로 미리보기 생성 완료. 필요 시 값 수정 후 생성하세요.")
            self.generate_btn.configure(state="normal")

            self._append_log("--- 권 자동생성 완료 ---")
            self._append_log(f"범위: {self.volume_spec_var.get().strip()}")
            self._append_log(f"생성 권수: {len(self.preview_records)}")
            self._append_log(f"작업일(일괄): {manual_date if manual_date else '(미입력)'}")
            self._append_log(f"지급명령번호(일괄): {manual_payment if manual_payment else '(미입력)'}")
        except Exception as exc:
            self.generate_btn.configure(state="disabled")
            self._append_log(f"Error: {exc}")
            messagebox.showerror("권 자동생성 실패", str(exc))

    def _validate_data(self) -> None:
        if self.worker_running:
            return

        excel_path = Path(self.excel_path_var.get().strip()).expanduser()
        output_dir = Path(self.output_dir_var.get().strip()).expanduser()

        try:
            if not excel_path.exists():
                raise FileNotFoundError(f"Excel 파일이 없습니다: {excel_path}")

            default_date = read_default_date_string(excel_path)
            records = read_cover_records(excel_path, default_date=default_date)

            self.preview_records = records
            self._refresh_table(self.preview_records)

            if default_date and not self.manual_date_var.get().strip():
                self.manual_date_var.set(default_date)

            self._refresh_summary_metrics()
            self.progress_var.set(0)
            self.progress_text_var.set(f"0 / {len(records)} (0%)")
            self.status_var.set("검증 완료. 필요하면 날짜/지급명령번호 입력 후 생성하세요.")

            self.generate_btn.configure(state="normal")

            self._append_log("--- 데이터 확인 완료 ---")
            self._append_log(f"Excel: {excel_path}")
            self._append_log(f"Output: {output_dir}")
            self._append_log(f"검출 권수: {len(records)}")
            if default_date:
                self._append_log(f"cover!{DATE_CELL} 기본 작업일: {default_date}")
        except Exception as exc:
            self.generate_btn.configure(state="disabled")
            self.status_var.set("검증 실패")
            self._append_log(f"Error: {exc}")
            messagebox.showerror("데이터 확인 실패", str(exc))

    def _apply_manual_to_preview(self) -> None:
        if not self.preview_records:
            messagebox.showwarning("미리보기 없음", "먼저 '데이터 확인'을 실행하세요.")
            return

        manual_date = self.manual_date_var.get().strip()
        manual_payment = self.manual_payment_var.get().strip()

        if not manual_date and not manual_payment:
            messagebox.showwarning("입력 필요", "작업일 또는 지급명령번호를 입력하세요.")
            return

        for rec in self.preview_records:
            if manual_date:
                rec.work_date = manual_date
            if manual_payment:
                rec.payment_no = manual_payment

        self._refresh_table(self.preview_records)
        self._refresh_summary_metrics()

        self._append_log("미리보기 일괄 반영 완료")
        self._append_log(f"작업일: {manual_date if manual_date else '(변경 없음)'}")
        self._append_log(f"지급명령번호: {manual_payment if manual_payment else '(변경 없음)'}")

    def _build_run_records(self) -> list[CoverRecord]:
        records = [CoverRecord(r.volume, r.work_date, r.payment_no) for r in self.preview_records]

        if self.apply_manual_var.get():
            manual_date = self.manual_date_var.get().strip()
            manual_payment = self.manual_payment_var.get().strip()
            if manual_date or manual_payment:
                for rec in records:
                    if manual_date:
                        rec.work_date = manual_date
                    if manual_payment:
                        rec.payment_no = manual_payment

                for rec in self.preview_records:
                    if manual_date:
                        rec.work_date = manual_date
                    if manual_payment:
                        rec.payment_no = manual_payment
                self._refresh_table(self.preview_records)
                self._refresh_summary_metrics()

                self._append_log("생성 직전 일괄 덮어쓰기 적용")

        return records

    def _start_generation(self) -> None:
        if self.worker_running:
            return

        if not self.preview_records:
            self._validate_data()
            if not self.preview_records:
                return

        run_records = self._build_run_records()
        output_dir = Path(self.output_dir_var.get().strip()).expanduser()

        if not any(r.work_date.strip() for r in run_records):
            messagebox.showerror("입력 필요", "작업일이 비어 있습니다. 작업일을 입력하거나 Excel 값을 확인하세요.")
            return

        self.cancel_event.clear()
        self.worker_running = True
        self.start_time = time.time()

        self.validate_btn.configure(state="disabled")
        self.generate_btn.configure(state="disabled")
        self.cancel_btn.configure(state="normal")

        self.excel_browse_btn.configure(state="disabled")
        self.output_browse_btn.configure(state="disabled")
        self.excel_entry.configure(state="disabled")
        self.output_entry.configure(state="disabled")

        self.manual_date_entry.configure(state="disabled")
        self.manual_payment_entry.configure(state="disabled")
        self.volume_spec_entry.configure(state="disabled")
        self.apply_manual_chk.configure(state="disabled")
        self.apply_manual_btn.configure(state="disabled")
        self.clipboard_load_btn.configure(state="disabled")
        self.range_build_btn.configure(state="disabled")

        self.done_var.set("0")
        self.failed_var.set("0")
        self.remaining_var.set(str(len(run_records)))
        self.current_var.set("-")
        self.progress_var.set(0)
        self.progress_text_var.set(f"0 / {len(run_records)} (0%)")
        self.status_var.set("생성 시작...")

        self._append_log("--- PDF 생성 시작 ---")

        self.worker_thread = threading.Thread(
            target=self._generate_worker,
            args=(run_records, output_dir),
            daemon=True,
        )
        self.worker_thread.start()
        self._tick_elapsed()

    def _cancel_generation(self) -> None:
        if not self.worker_running:
            return
        self.cancel_event.set()
        self.status_var.set("중단 요청됨... 현재 권 처리 후 멈춥니다.")
        self._append_log("중단 요청 수신")
        self.cancel_btn.configure(state="disabled")

    def _generate_worker(self, records: list[CoverRecord], output_dir: Path) -> None:
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            total = len(records)
            success_count = 0
            fail_count = 0

            for idx, rec in enumerate(records, start=1):
                if self.cancel_event.is_set():
                    self.log_queue.put(("cancelled", (success_count, fail_count, total)))
                    return

                self.log_queue.put(("current", rec.volume))
                self.log_queue.put(("item_status", (idx - 1, "생성중", "", "running")))
                self.log_queue.put(("status", f"[{idx}/{total}] {rec.volume}권 생성 중..."))
                self.log_queue.put(("log", f"Generating volume {rec.volume} ({idx}/{total})"))

                try:
                    saved_path = generate_cover_pdf(rec, output_dir)
                    success_count += 1
                    self.log_queue.put(("item_status", (idx - 1, "완료", str(saved_path), "done")))
                    self.log_queue.put(("log", f"Saved: {saved_path}"))
                except Exception as exc:
                    fail_count += 1
                    self.log_queue.put(("item_status", (idx - 1, "실패", str(exc), "failed")))
                    self.log_queue.put(("log", f"Failed volume {rec.volume}: {exc}"))

                self.log_queue.put(("counts", (success_count, fail_count, total)))
                self.log_queue.put(("progress", (idx, total)))

            self.log_queue.put(("done", (success_count, fail_count, total)))
        except Exception as exc:
            self.log_queue.put(("error", str(exc)))

    def _poll_queue(self) -> None:
        try:
            while True:
                event, payload = self.log_queue.get_nowait()

                if event == "log":
                    self._append_log(str(payload))
                elif event == "status":
                    self.status_var.set(str(payload))
                elif event == "progress":
                    current, total = payload
                    ratio = 0 if total == 0 else (current / total) * 100
                    self.progress_var.set(ratio)
                    self.progress_text_var.set(f"{current} / {total} ({ratio:.1f}%)")
                elif event == "counts":
                    success_count, fail_count, total = payload
                    self.done_var.set(str(success_count))
                    self.failed_var.set(str(fail_count))
                    self.remaining_var.set(str(max(total - success_count - fail_count, 0)))
                elif event == "current":
                    self.current_var.set(str(payload))
                elif event == "item_status":
                    idx, status_text, detail_text, tag = payload
                    item_id = f"vol_{idx}"
                    if self.volume_table.exists(item_id):
                        no, volume, work_date, payment_no, _, _ = self.volume_table.item(item_id, "values")
                        self.volume_table.item(
                            item_id,
                            values=(no, volume, work_date, payment_no, status_text, detail_text),
                            tags=(tag,),
                        )
                        self.volume_table.see(item_id)
                elif event == "done":
                    success_count, fail_count, total = payload
                    self.status_var.set(f"완료: 성공 {success_count}, 실패 {fail_count}, 전체 {total}")
                    self.progress_var.set(100)
                    self._append_log("--- 생성 완료 ---")
                    self._end_worker_state()
                    messagebox.showinfo(
                        "완료",
                        f"PDF 생성이 완료되었습니다.\n\n성공: {success_count}\n실패: {fail_count}\n전체: {total}",
                    )
                elif event == "cancelled":
                    success_count, fail_count, total = payload
                    self.status_var.set(f"중단됨: 성공 {success_count}, 실패 {fail_count}, 전체 {total}")
                    self._append_log("--- 작업 중단 ---")
                    self._end_worker_state()
                    messagebox.showwarning(
                        "중단됨",
                        f"작업이 중단되었습니다.\n\n성공: {success_count}\n실패: {fail_count}\n전체: {total}",
                    )
                elif event == "error":
                    self.status_var.set("실패")
                    self._append_log(f"Error: {payload}")
                    self._end_worker_state()
                    messagebox.showerror("생성 실패", str(payload))
        except Empty:
            pass

        self.root.after(120, self._poll_queue)

    def _refresh_summary_metrics(self) -> None:
        total = len(self.preview_records)
        self.total_var.set(str(total))
        self.done_var.set("0")
        self.failed_var.set("0")
        self.remaining_var.set(str(total))
        self.current_var.set("-")

        dates = {r.work_date.strip() for r in self.preview_records if r.work_date.strip()}
        if not dates:
            self.date_var.set("-")
        elif len(dates) == 1:
            self.date_var.set(next(iter(dates)))
        else:
            self.date_var.set(f"복수({len(dates)})")

    def _end_worker_state(self) -> None:
        self.worker_running = False
        self.cancel_event.clear()

        self.validate_btn.configure(state="normal")
        self.generate_btn.configure(state="normal" if self.preview_records else "disabled")
        self.cancel_btn.configure(state="disabled")

        self.excel_browse_btn.configure(state="normal")
        self.output_browse_btn.configure(state="normal")
        self.excel_entry.configure(state="normal")
        self.output_entry.configure(state="normal")

        self.manual_date_entry.configure(state="normal")
        self.manual_payment_entry.configure(state="normal")
        self.volume_spec_entry.configure(state="normal")
        self.apply_manual_chk.configure(state="normal")
        self.apply_manual_btn.configure(state="normal")
        self.clipboard_load_btn.configure(state="normal")
        self.range_build_btn.configure(state="normal")

        self.current_var.set("-")

    def _tick_elapsed(self) -> None:
        if self.worker_running and self.start_time is not None:
            elapsed = max(int(time.time() - self.start_time), 0)
            minutes = elapsed // 60
            seconds = elapsed % 60
            self.elapsed_var.set(f"{minutes:02d}:{seconds:02d}")
            self.root.after(500, self._tick_elapsed)

    def _refresh_table(self, records: list[CoverRecord]) -> None:
        for item in self.volume_table.get_children():
            self.volume_table.delete(item)

        for idx, rec in enumerate(records, start=1):
            self.volume_table.insert(
                "",
                "end",
                iid=f"vol_{idx - 1}",
                values=(idx, rec.volume, rec.work_date, rec.payment_no, "대기", ""),
                tags=("waiting",),
            )

    def _append_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        line = f"[{timestamp}] {message}"
        self.log_text.configure(state="normal")
        self.log_text.insert("end", line + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")


def main() -> None:
    root = tk.Tk()
    CoverGeneratorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
